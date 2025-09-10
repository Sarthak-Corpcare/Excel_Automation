import openpyxl
from datetime import datetime , timedelta
import sys
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# --- CONFIGURATION ---
raw_file = "Daily Performance - 10th September 2025.xlsx"
template_file = "Daily Performance Sheet - 05 Aug 2025.xlsx"
output_file = "Daily Performance - Filled.xlsx"
SHEETS_TO_IGNORE = ['Home', 'Sheet1', 'Disclaimer']
LOGO_FILENAME = "corpcare_logo.jpg"

# --- MAPPING & STYLING CONFIGURATION ---
RAW_TO_TEMPLATE_HEADER_MAP = {
    "Scheme Name": "Scheme Name",  "Average Maturity Years": "Average Maturity Years",
    "Modified Duration Years": "Modified Duration Years", "YTM (%)": "YTM (%)",
    "Direct Expense Ratio": "Direct Expense Ratio",
    "Latest Date": "Latest Date", "Latest NAV(Rs)": "Latest NAV(Rs)", "1 Day": "1 Day", "3 Day": "3 Day",
    "1 Week": "1 Week",
    "2 Week": "2 Week", "1 Month": "1 Month", "3 Months": "3 Months", "6 Months": "6 Months", "9 Months": "9 Months",
    "1 Year": "1 Year", "3 Years": "3 Years", "5 Years": "5 Years", "10 Years": "10 Years",
    "SINCE INCEPTION": "SINCE INCEPTION",
    "Cash & Equivalent": "Cash & Equivalent", "Others": "Others", "SOV": "SOV",
    "A1 / A1+ / A1-": "A1 / A1+ / A1-",
    "A / A+ / A1+ / A1-": "A / A+ / A1+ / A1-",
    "AA / AA+ / AA-": "AA / AA+ / AA-", "AAA": "AAA", "Unrated": "Unrated", "D": "D", "A1+ / A1-": "A1+ / A1-",
    "Exit Load": "Exit Load", "Remark": "Remark", "Inception Date": "Inception Date",
    "[Fund Manager 1]": "[Fund Manager 1]",
}
MONTH_GROUPED_HEADERS = ["Average Maturity Years", "Modified Duration Years", "YTM (%)", "Direct Expense Ratio",
                         "Cash & Equivalent", "Others", "SOV", "A / A+ / A1+ / A1-", "AA / AA+ / AA-", "AAA", "Unrated",
                         "D", "A1+ / A1-"]

# --- GLOBAL STYLES ---
light_brown_fill = PatternFill(start_color="DCC7A3", end_color="DCC7A3", fill_type="solid")
no_fill = PatternFill(fill_type=None)
back_button_fill = PatternFill(start_color="DCC783", end_color="DCC783", fill_type="solid")
back_button_font = Font(name='Calibri', size=11, color="000000", bold=True, underline=None)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border_side = Side(border_style="thin", color="000000")
button_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)


# --- HELPER FUNCTIONS ---
def is_meaningful_data(val):
    """Checks if a cell contains actual data, not just blanks, zeros, or hyphens."""
    if val is None:
        return False
    if isinstance(val, str) and val.strip() in ["", "-"]:
        return False
    if isinstance(val, (int, float)) and val == 0:
        return False
    return True


def is_date_like(v):
    if v is None: return None
    if isinstance(v, datetime): return v
    if isinstance(v, str):
        for fmt in ("%d-%b-%Y", "%d-%b-%y", "%d-%b-%Y ", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(v.strip(), fmt)
            except Exception:
                continue
    return None


def find_header_row(sheet, keyword="Scheme Name"):
    for r in range(1, 25):
        for cell in sheet[r]:
            try:
                if cell.value and str(cell.value).strip() == keyword: return r
            except Exception:
                continue
    return -1


def update_as_on_date(sheet):
    for r in range(1, 11):
        for cell in sheet[r]:
            if cell.value and str(cell.value).strip().startswith("As on"):
                yesterday = datetime.today() - timedelta(days=1)

                cell.value = f"As on {yesterday.strftime('%Y-%b-%d')}"
                return True
    return False


def get_month_id_for_column(sheet, row, col):
    for r in range(row - 1, 0, -1):
        cell = sheet.cell(row=r, column=col)
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= r <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                if isinstance(top_left_cell.value, int) and len(str(top_left_cell.value)) == 6:
                    return top_left_cell.value
    return None




def get_parent_header_for_column(sheet, header_row, col):
    for r in range(header_row - 1, 0, -1):
        cell = sheet.cell(row=r, column=col)
        val = cell.value
        if val is None:
            for merged_range in sheet.merged_cells.ranges:
                if merged_range.min_row <= r <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                    top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    if top_left.value: return str(top_left.value).strip()
            continue
        if isinstance(val, int) and len(str(val)) == 6: continue
        if is_date_like(val): continue
        return str(val).strip()
    return None


def create_styled_homepage(workbook):

    print("Creating precise styled homepage...")
    if 'Home' in workbook.sheetnames :

        home_sheet = workbook['Home']
        home_sheet.delete_rows(1, home_sheet.max_row + 1)
        while home_sheet.merged_cells.ranges:
            home_sheet.unmerge_cells(str(home_sheet.merged_cells.ranges[0]))
    else:
        home_sheet = workbook.create_sheet('Home', 0)

    # Define styles locally for this function
    title_font = Font(name='Calibri', size=14, bold=True)
    date_font = Font(name='Calibri', size=11, color="003366")

    # *** THIS IS THE ONLY CHANGE: Corrected font for the homepage buttons ***
    button_font_ui = Font(name='Calibri', size=11, color="000000", bold=True, underline=None)

    heading_fill = PatternFill(start_color="D1B27B", end_color="D1B27B", fill_type="solid")
    button_fill_ui = PatternFill(start_color="DCC783", end_color="DCC783", fill_type="solid")
    right_align = Alignment(horizontal='right', vertical='center')
    img = None
    try:
        img = Image(LOGO_FILENAME)
        img.height = 105
        img.width = (img.width / img.height) * img.height
        home_sheet.add_image(img, 'A1')
        text_cell = home_sheet['A7']
        text_cell.value = "CorpCare Investment Advisory Pvt Ltd."
        text_cell = home_sheet['A8']
        text_cell.value = "RIA : INA000018249"
    except FileNotFoundError:
        print(f"    WARNING: Logo file '{LOGO_FILENAME}' not found.")

    home_sheet.merge_cells('H1:L2')
    title_cell = home_sheet['H1']
    title_cell.value = "Daily Debt MF Tracker"
    title_cell.font = title_font
    title_cell.fill = heading_fill
    title_cell.alignment = center_align
    for row in home_sheet['H1:L2']:
        for cell in row: cell.border = button_border

    date_cell = home_sheet['O2']
    date_cell.value = datetime.today().strftime('%d-%b-%y')
    date_cell.font = date_font
    date_cell.alignment = right_align

    home_sheet.merge_cells('E4:O5')
    debt_funds_cell = home_sheet['E4']
    debt_funds_cell.value = "Debt Funds"
    debt_funds_cell.font = title_font
    debt_funds_cell.fill = heading_fill
    debt_funds_cell.alignment = center_align
    for row in home_sheet['E4:O5']:
        for cell in row: cell.border = button_border

    data_sheets = sorted([s for s in workbook.sheetnames if s not in SHEETS_TO_IGNORE and s.strip().lower() != 'disclaimer'])
    start_row, start_col, max_cols = 7, 5, 4
    button_height, button_width, row_gap, col_gap = 2, 2, 1, 1
    for i, sheet_name in enumerate(data_sheets):
        row_index = i // max_cols
        col_index = i % max_cols
        cell_row = start_row + (row_index * (button_height + row_gap))
        cell_col = start_col + (col_index * (button_width + col_gap))
        home_sheet.merge_cells(start_row=cell_row, end_row=cell_row + button_height - 1, start_column=cell_col,
                               end_column=cell_col + button_width - 1)
        button_cell = home_sheet.cell(row=cell_row, column=cell_col)
        button_cell.value = sheet_name
        button_cell.alignment = center_align
        for r_offset in range(button_height):
            for c_offset in range(button_width):
                cell_to_style = home_sheet.cell(row=cell_row + r_offset, column=cell_col + c_offset)
                cell_to_style.fill = button_fill_ui
                cell_to_style.border = button_border
        button_cell.hyperlink = f"#'{sheet_name}'!A1"
        button_cell.font = button_font_ui

    home_sheet.sheet_view.showGridLines = False
    if img: home_sheet.column_dimensions['A'].width = (img.width / 7)
    for i in range(2, 20): home_sheet.column_dimensions[get_column_letter(i)].width = 12
    print("Styled homepage created successfully.")


# --- MAIN SCRIPT ---
print("Starting Data Transfer")
try:
    raw_wb = openpyxl.load_workbook(raw_file, data_only=True)
    template_wb = openpyxl.load_workbook(template_file)
except FileNotFoundError as e:
    print(f"ERROR: Could not find a required file: {e.filename}")
    sys.exit()

print("Processing and writing data one sheet at a time")
total_rows_written = 0
for sheet_name in raw_wb.sheetnames:
    if sheet_name in SHEETS_TO_IGNORE: continue
    print(f"--- Processing sheet: '{sheet_name}' ---")
    raw_sheet = raw_wb[sheet_name]
    if sheet_name not in template_wb.sheetnames:
        print(f"WARNING: Sheet '{sheet_name}' not found in template file. Skipping.")
        continue
    template_sheet = template_wb[sheet_name]

    update_as_on_date(template_sheet)
    data_header_row_raw = find_header_row(raw_sheet)
    if data_header_row_raw == -1: continue
    data_header_row_template = find_header_row(template_sheet)
    if data_header_row_template == -1: continue

    # --- SOURCE COLUMNS ---
    header_row_str = str(data_header_row_raw)

    # Hardcode the latest AUM to column C and the older AUM to column B.
    latest_aum_col_raw = 3  # Column C
    older_aum_col_raw = 2  # Column B

    # Get the date for the legend directly from the known cell B12.
    older_date_for_legend = raw_sheet['B' + header_row_str].value

    raw_col_map = {c: {'header': str(cell.value).strip() if cell.value else "",
                       'month_id': get_month_id_for_column(raw_sheet, data_header_row_raw, c)} for c, cell in
                   enumerate(raw_sheet[data_header_row_raw], 1)}
    all_month_ids = sorted(list(set(v['month_id'] for v in raw_col_map.values() if v['month_id'] is not None)),
                           reverse=True)
    latest_month_id = all_month_ids[0] if all_month_ids else None
    older_month_id = all_month_ids[1] if len(all_month_ids) > 1 else None

    raw_date_columns_by_parent = {}
    for col, cell in enumerate(raw_sheet[data_header_row_raw], 1):
        if (date_v := is_date_like(cell.value)):
            parent = get_parent_header_for_column(raw_sheet, data_header_row_raw, col) or "(unknown parent)"
            raw_date_columns_by_parent.setdefault(parent, []).append((col, date_v))
    for p, lst in raw_date_columns_by_parent.items():
        raw_date_columns_by_parent[p] = sorted(lst, key=lambda t: t[1], reverse=True)

    # --- DESTINATION COLUMNS ---
    dest_col_map = {str(c.value).strip(): c.column for c in template_sheet[data_header_row_template] if c.value}
    aum_dest_col = next(
        (c.column for r in range(data_header_row_template - 2, data_header_row_template + 1) for c in template_sheet[r]
         if "AUM" in str(c.value)), None)

    # --- TEMPLATE SHEET ---
    if (benchmark_row := next((r for r in range(1, template_sheet.max_row + 2) if
                               str(template_sheet.cell(row=r, column=1).value).strip().lower() == "benchmark"), None)):
        template_sheet.delete_rows(benchmark_row, template_sheet.max_row - benchmark_row + 2)
        print("    Benchmark table removed.")
    clear_from = data_header_row_template + 1
    if template_sheet.max_row >= clear_from:
        for row in template_sheet.iter_rows(min_row=clear_from):
            if not row[0].value or str(row[0].value).strip() == "": break
            for cell in row: cell.value = None

    # --- WRITING DATA ROW BY ROW ---
    start_row_template = data_header_row_template + 1
    rows_on_this_sheet = 0
    any_older_data_used = False
    exp_dates = next((v for k, v in raw_date_columns_by_parent.items() if 'expense' in k.lower()), [])

    for row_num in range(data_header_row_raw + 1, raw_sheet.max_row + 2):
        if not raw_sheet.cell(row=row_num, column=1).value: break
        template_row_index = start_row_template + rows_on_this_sheet

        # A) Process all standard columns
        # --- inside row loop ---

        # A1) Handle standard columns (non-rating)
        for raw_h, tpl_h in RAW_TO_TEMPLATE_HEADER_MAP.items():
            if tpl_h in dest_col_map and tpl_h not in ["AAA", "AA / AA+ / AA-", "A / A+ / A1+ / A1-", "D", "Unrated",
                                                       "Cash & Equivalent", "Others", "SOV"]:
                dest_col = dest_col_map[tpl_h]
                final_value, is_older = None, False
                latest_col = next(
                    (c for c, v in raw_col_map.items() if v['header'] == raw_h and v['month_id'] == latest_month_id),
                    None)
                older_col = next(
                    (c for c, v in raw_col_map.items() if v['header'] == raw_h and v['month_id'] == older_month_id),
                    None)
                static_col = next((c for c, v in raw_col_map.items() if v['header'] == raw_h and v['month_id'] is None),
                                  None)

                if latest_col:
                    latest_val = raw_sheet.cell(row=row_num, column=latest_col).value
                    if is_meaningful_data(latest_val):
                        final_value = latest_val

                if final_value is None and older_col:
                    older_val = raw_sheet.cell(row=row_num, column=older_col).value
                    if is_meaningful_data(older_val):
                        final_value = older_val
                        is_older = True
                        any_older_data_used = True

                if final_value is None and static_col:
                    final_value = raw_sheet.cell(row=row_num, column=static_col).value

                dest_cell = template_sheet.cell(row=template_row_index, column=dest_col)
                dest_cell.value = final_value
                if tpl_h in MONTH_GROUPED_HEADERS:
                    dest_cell.fill = light_brown_fill if is_older else no_fill

        # A2) Handle rating allocation block (special rule)
        RATING_HEADERS = ["A / A+ / A-", "AA / AA+ / AA-", "A / A+ / A1+ / A1-","AAA","A1 / A1+ / A1-", "Cash & Equivalent","D","Others","SOV", "Unrated"]

        # Step 1: check if latest has *any* rating filled
        latest_has_any = False
        latest_vals = {}
        for raw_h in RATING_HEADERS:
            latest_col = next(
                (c for c, v in raw_col_map.items() if v['header'] == raw_h and v['month_id'] == latest_month_id), None)
            if latest_col:
                val = raw_sheet.cell(row=row_num, column=latest_col).value
                if is_meaningful_data(val):
                    latest_has_any = True
                latest_vals[raw_h] = val

        # Step 2: choose source (latest vs older)
        use_latest = latest_has_any
        for raw_h in RATING_HEADERS:
            tpl_h = raw_h
            if tpl_h not in dest_col_map:
                continue
            dest_col = dest_col_map[tpl_h]

            if use_latest:
                final_val = latest_vals.get(raw_h)
                is_older = False
            else:
                older_col = next(
                    (c for c, v in raw_col_map.items() if v['header'] == raw_h and v['month_id'] == older_month_id),
                    None)
                final_val, is_older = None, False
                if older_col:
                    val = raw_sheet.cell(row=row_num, column=older_col).value
                    if is_meaningful_data(val):
                        final_val = val
                        is_older = True
                        any_older_data_used = True

            dest_cell = template_sheet.cell(row=template_row_index, column=dest_col)
            dest_cell.value = final_val
            dest_cell.fill = light_brown_fill if is_older else no_fill

        # B) AUM
        if aum_dest_col:
            aum_value, aum_is_older = None, False
            # print(f"      - AUM Processing:")

            # Check Latest
            if latest_aum_col_raw:
                val = raw_sheet.cell(row=row_num, column=latest_aum_col_raw).value
                # print(f"        > Checking LATEST AUM (col {latest_aum_col_raw})... Found: '{val}'")
                if is_meaningful_data(val):
                    aum_value = val

            # Check Older (if needed)
            if aum_value is None:

                if older_aum_col_raw:
                    val = raw_sheet.cell(row=row_num, column=older_aum_col_raw).value
                    print(f"        > LATEST was empty. Checking OLDER AUM (col {older_aum_col_raw})... Found: '{val}'")

                    # print(all_raw_aum_cells, len(all_raw_aum_cells))
                    print(latest_aum_col_raw)
                    print(older_aum_col_raw)
                    if is_meaningful_data(val):
                        aum_value = val
                        aum_is_older = True
                        any_older_data_used = True
                # else:
                    # print(f"        > LATEST was empty and NO OLDER AUM column to check.")

            dest_cell = template_sheet.cell(row=template_row_index, column=aum_dest_col)
            dest_cell.value = aum_value
            dest_cell.fill = light_brown_fill if aum_is_older else no_fill
            # print(
            #     f"        -> FINAL AUM ACTION: Wrote value '{aum_value}'. Coloring: {'YES' if aum_is_older else 'NO'}.")

        # C) Historical Expense Ratio
        if exp_dates and (dest_col := dest_col_map.get("Direct Expense Ratio")):
            her_value, her_is_older = None, False
            if (val := raw_sheet.cell(row=row_num, column=exp_dates[0][0]).value) is not None and str(
                    val).strip() != "":
                her_value = val
            if her_value is None and len(exp_dates) > 1 and (
            val := raw_sheet.cell(row=row_num, column=exp_dates[1][0]).value) is not None and str(val).strip() != "":
                her_value = val
                her_is_older = True
                any_older_data_used = True
            dest_cell = template_sheet.cell(row=template_row_index, column=dest_col)
            dest_cell.value = her_value
            dest_cell.fill = light_brown_fill if her_is_older else no_fill

        rows_on_this_sheet += 1
        total_rows_written += 1
    print(f" Wrote {rows_on_this_sheet} rows of new data.")



    # --- FORMAT & LEGEND ---
    if aum_dest_col:
        try:
            template_sheet.merge_cells(start_row=data_header_row_template - 2, end_row=data_header_row_template - 1,
                                       start_column=aum_dest_col, end_column=aum_dest_col)
            corpus_cell = template_sheet.cell(row=data_header_row_template - 2, column=aum_dest_col)
            corpus_cell.value = "Corpus"
            corpus_cell.alignment = center_align
            template_sheet.cell(row=data_header_row_template, column=aum_dest_col).value = "AUM (Cr.)"
        except Exception as e:
            print(f" Warning: could not merge Corpus header: {e}")
    if any_older_data_used:
        older_date_str = older_date_for_legend.strftime('%d-%b-%Y') if older_date_for_legend else (
                    older_month_id and datetime.strptime(str(older_month_id), "%Y%m").strftime("%b-%Y"))
        if older_date_str:
            legend_row = start_row_template + rows_on_this_sheet + 2
            template_sheet.cell(row=legend_row, column=1).fill = light_brown_fill
            template_sheet.column_dimensions['A'].width = 5
            template_sheet.cell(row=legend_row, column=2).value = f"Indicates data as of {older_date_str}"
            template_sheet.cell(row=legend_row, column=2).font = Font(bold=True)

    # --- HOME BUTTON ---
    back_button = template_sheet['A2']
    back_button.value = "Home"
    back_button.hyperlink = f"#'Home'!A1"
    back_button.font = back_button_font
    back_button.fill = back_button_fill
    back_button.alignment = center_align
    back_button.border = button_border
    template_sheet.column_dimensions['A'].width = 50

# --- FINAL HOMEPAGE ---
create_styled_homepage(template_wb)

print(f"\nTotal rows written across all sheets: {total_rows_written}")
template_wb.save(output_file)
print(f"Success! Data transferred and saved to {output_file}")
